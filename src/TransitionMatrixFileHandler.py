import numpy as np
import os
import math

from openpyxl import Workbook, load_workbook

from TransitionMatrix import TransitionMatrix
from States import STATES

def save_matrix_to_txt(tm: TransitionMatrix, file):
    """Guarda una TransitionMatrix a un archivo de texto.

    Arguments:
        tm -- La TransitionMatrix a guardar
        file -- El path del archivo en el que se guarda la TransitionMatrix.
    """    
    np.savetxt(file, tm.matrix)

def save_matrix_to_xlsx(tm: TransitionMatrix, file):
    """Guarda una TransitionMatrix a un archivo xlsx.

    Arguments:
        tm -- La TransitionMatrix a guardar
        file -- El path del archivo en el que se guarda la TransitionMatrix
    """    

    workbook = Workbook()
    sheet = workbook.active

    for i in range(tm.matrix.shape[0]):
        row = tm.matrix[i,:].tolist()
        sheet.append(row)

    workbook.save(file)


def load_matrix_from_txt(file) -> TransitionMatrix:
    """Carga una TransitionMatrix desde un archivo de texto y la valida.

    Arguments:
        file -- El path del archivo de texto que contiene la TransitionMatrix

    Returns:
        La TransitionMatrix leída del archivo de texto, validada.
    """
    matrix = np.loadtxt(file)

    check_if_matrix_is_valid(matrix)

    return TransitionMatrix(matrix)

def load_matrix_from_xlsx(file) -> TransitionMatrix:
    """Carga una TransitionMatrix desde un archivo xlsx y la valida.

    Arguments:
        file -- El path del archivo xlsx que contiene la TransitionMatrix

    Returns:
        La TransitionMatrix leída del archivo xlsx, validada.
    """    

    workbook = load_workbook(filename=file)
    sheet = workbook.active
    rows = []

    for r in sheet.values:
        rows.append(r)

    matrix = np.array(rows)

    check_if_matrix_is_valid(matrix)

    return TransitionMatrix(matrix)



def check_if_matrix_is_valid(matrix: np.ndarray):
    """Revisa si una matriz cumple con las condiciones para ser una TransitionMatrix válida.

    Las condiciones son dos:
    - Los valores de cada fila deben sumar 1.
    - La matriz debe tener una número apropiado de filas y columnas, correspondientes al número de estados con los que se está trabajando.

    Arguments:
        matrix -- La matriz a validar.

    Raises:
        Exception: La forma de la matriz no es consistente con el número de estados
        Exception: Al menos una fila de la matriz tiene una suma distinta de 1.
    """    
    if matrix.shape != (len(STATES), len(STATES)):
        raise Exception("La matriz tiene una forma no válida ({})".format(matrix.shape))
    
    for i in range(len(STATES)):
        row = matrix[i, :]
        row_total = sum(row)
        if not math.isclose(row_total, 1):
            raise Exception("La fila {} suma {} en lugar de 1.".format(i, row_total))
    
def save_results_to_file(col: np.ndarray, file, format='list'):
    """Guarda un np.ndrray, que contiene resultados de procesar una TransitionMatrix, en un archivo de texto.

    Arguments:
        col -- Los resultados que se quieren guardar
        file -- El archivo en el que se guardarán los reusltados
        format -- El formato que tendrá la información dentro del archivo

    Raises:
        Exception: El formato especificado debe ser uno de los formatos válidos.
    """    

    filename, fileext = os.path.splitext(file)
    if fileext == '.xlsx':
        array_to_xlsx(col, file)

    else:
        if format not in ('list, dict'):
            raise Exception("'{}' no es un formato válido.")
    
        if format == 'list':
            np.savetxt(file, col)

        elif format == 'dict':
            probs_dict = dict(zip(STATES.keys(), col))
            with open(file, 'w') as f:
                print('{', file=f)
                for item in probs_dict.items():
                    print(item, file=f)
                print('}', file=f)

def array_to_xlsx(arr: np.ndarray, file):
    workbook = Workbook()
    sheet = workbook.active

    arr_list = arr.tolist()
    sheet.append(arr_list)
    workbook.save(file)